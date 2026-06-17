import io
import json
import logging
import re
from datetime import datetime

from flask import (
    Blueprint, jsonify, redirect, render_template,
    request, url_for, flash,
)
from werkzeug.utils import secure_filename

log = logging.getLogger(__name__)

from extensions import db
from models import AppSettings, ComparisonJob, Competition, Edition, Document
from services.comparator import (
    compare_one_pair,
    generate_edition_summary_text,
    run_comparison,
    get_pair_structures,
    compare_sections_batch,
    generate_pair_summary,
    make_gemini_caller,
    BATCH_SIZE,
)

bp = Blueprint("comparison", __name__, url_prefix="/comparison")

PRICING = {
    "gemini-2.5-flash":      {"input": 0.30,  "output": 2.50},
    "gemini-2.5-pro":        {"input": 1.25,  "output": 10.00},
    "gemini-2.5-flash-lite": {"input": 0.10,  "output": 0.40},
}


@bp.route("/")
def index():
    jobs = ComparisonJob.query.order_by(ComparisonJob.created_at.desc()).all()
    return render_template("comparison/index.html", jobs=jobs)


@bp.route("/setup/", methods=["GET", "POST"])
def setup():
    settings = db.session.get(AppSettings, 1)
    if not settings or not settings.gemini_api_key:
        flash("Skonfiguruj klucz Gemini API w Ustawieniach.", "warning")
        return redirect(url_for("settings.index"))

    competitions = Competition.query.order_by(Competition.name).all()

    if request.method == "POST":
        competition_id = request.form.get("competition_id", type=int)
        edition_old_id = request.form.get("edition_old_id", type=int)
        edition_new_id = request.form.get("edition_new_id", type=int)

        if not competition_id or not edition_old_id or not edition_new_id:
            flash("Wybierz konkurs i dwie edycje.", "error")
            return redirect(url_for("comparison.setup"))
        if edition_old_id == edition_new_id:
            flash("Wybierz dwie różne edycje.", "error")
            return redirect(url_for("comparison.setup"))

        edition_old  = db.session.get(Edition, edition_old_id)
        edition_new  = db.session.get(Edition, edition_new_id)
        competition  = db.session.get(Competition, competition_id)

        mappings = []
        i = 0
        while True:
            old_id_str = request.form.get(f"mapping_{i}_old")
            if old_id_str is None:
                break
            new_id_str = request.form.get(f"mapping_{i}_new", "")
            if old_id_str and new_id_str and new_id_str != "__skip__":
                old_doc = db.session.get(Document, int(old_id_str))
                new_doc = db.session.get(Document, int(new_id_str))
                if old_doc and new_doc:
                    mappings.append({
                        "old_doc_id": int(old_id_str),
                        "new_doc_id": int(new_id_str),
                        "old_name":   old_doc.original_name,
                        "new_name":   new_doc.original_name,
                    })
            i += 1

        if not mappings:
            flash("Nie wybrano żadnych par plików do porównania.", "error")
            return redirect(url_for("comparison.setup"))

        active_statuses = ["pending", "comparing", "extracting", "chunking", "summarizing"]
        is_busy = ComparisonJob.query.filter(ComparisonJob.status.in_(active_statuses)).first() is not None
        initial_status = "queued" if is_busy else "pending"

        job = ComparisonJob(
            competition_name   = competition.name if competition else "",
            edition_old_id     = edition_old_id,
            edition_new_id     = edition_new_id,
            label_old          = edition_old.name if edition_old else "Edycja starsza",
            label_new          = edition_new.name if edition_new else "Edycja nowsza",
            file_mappings_json = json.dumps(mappings, ensure_ascii=False),
            status             = initial_status,
            progress_total     = len(mappings),
            progress_current   = 0,
            gemini_model_used  = settings.gemini_model,
            skip_redactional   = request.form.get("skip_redactional") == "on",
            job_label          = request.form.get("job_label", "").strip() or None,
        )
        db.session.add(job)
        db.session.commit()

        return redirect(url_for("comparison.job_status", job_id=job.id))

    return render_template("comparison/setup.html", competitions=competitions, settings=settings)


# ── AJAX endpoints for browser-driven comparison ───────────────────────────

@bp.route("/job/<int:job_id>/run-pair", methods=["POST"])
def run_pair(job_id):
    """Process one file pair synchronously in the main request thread."""
    job = ComparisonJob.query.get_or_404(job_id)

    # Catch cancellations that arrived while a previous pair was running
    db.session.refresh(job)
    if job.status == "cancelled":
        return jsonify({"ok": False, "cancelled": True, "error": "Porównanie zostało anulowane"})

    settings = db.session.get(AppSettings, 1)

    if not settings or not settings.gemini_api_key:
        return jsonify({"ok": False, "error": "Brak klucza Gemini API w ustawieniach"}), 400

    data = request.get_json() or {}
    pair_idx = data.get("pair_idx", 0)

    mappings = json.loads(job.file_mappings_json or "[]")
    if pair_idx >= len(mappings):
        return jsonify({"ok": False, "error": f"Nieprawidłowy indeks pary: {pair_idx}"}), 400

    mapping = mappings[pair_idx]
    doc_old = db.session.get(Document, mapping["old_doc_id"])
    doc_new = db.session.get(Document, mapping["new_doc_id"])

    if not doc_old or not doc_new:
        return jsonify({"ok": False, "pair_idx": pair_idx, "error": "Dokument nie istnieje"}), 404

    job.status = "comparing"
    job.status_detail = f"Para {pair_idx + 1}/{len(mappings)}: {doc_old.original_name}"
    job.progress_current = pair_idx
    if not job.started_at:
        job.started_at = datetime.utcnow()
    db.session.commit()

    pair_prefix = f"Para {pair_idx + 1}/{len(mappings)}"

    def _on_status(stage):
        job.status_detail = f"{pair_prefix}: {stage}"
        db.session.commit()

    log.debug("run_pair START  job=%d  para=%d/%d  stary=%s  nowy=%s",
              job_id, pair_idx + 1, len(mappings),
              doc_old.original_name, doc_new.original_name)

    try:
        result = compare_one_pair(doc_old, doc_new, job, settings, on_status=_on_status)
        result.setdefault("compared_at", datetime.utcnow().strftime("%d.%m.%Y %H:%M"))

        per_file_results = json.loads(job.per_file_results_json or "[]")
        per_file_results = [r for r in per_file_results if r.get("pair_idx") != pair_idx]
        per_file_results.append(result)
        job.per_file_results_json = json.dumps(per_file_results, ensure_ascii=False)

        all_changes = [c for r in per_file_results for c in r.get("changes", [])]
        job.changes_json     = json.dumps(all_changes, ensure_ascii=False)
        job.progress_current = pair_idx + 1
        job.tokens_input     = (job.tokens_input  or 0) + result.get("tokens_in",  0)
        job.tokens_output    = (job.tokens_output or 0) + result.get("tokens_out", 0)
        job.error_message    = None  # Wyczyść poprzedni błąd jeśli ponowienie się udało
        _update_cost(job)
        db.session.commit()

        log.debug("run_pair OK  job=%d  para=%d  zmian=%d",
                  job_id, pair_idx + 1, len(result.get("changes", [])))

        import markdown as md_lib
        summary_html = md_lib.markdown(result.get("summary", ""), extensions=["extra"]) if result.get("summary") else ""

        return jsonify({
            "ok":          True,
            "pair_idx":    pair_idx,
            "old_name":    result["old_name"],
            "new_name":    result["new_name"],
            "changes":     result["changes"],
            "summary":     result.get("summary", ""),
            "summary_html": summary_html,
        })
    except Exception as exc:
        log.error("run_pair BŁĄD  job=%d  para=%d  %s: %s",
                  job_id, pair_idx + 1, type(exc).__name__, exc, exc_info=True)
        try:
            job.status        = "error"
            job.error_message = str(exc)[:1000]
            db.session.commit()
        except Exception as db_err:
            log.error("run_pair DB commit nieudany: %s", db_err)
            try:
                db.session.rollback()
            except Exception:
                pass
        return jsonify({"ok": False, "pair_idx": pair_idx, "error": str(exc)}), 500


@bp.route("/job/<int:job_id>/finish-pairs", methods=["POST"])
def finish_pairs(job_id):
    """Mark all pairs as processed — sets status to awaiting_summary."""
    job = ComparisonJob.query.get_or_404(job_id)
    if job.status in ("pending", "comparing", "error"):
        per_file_results = json.loads(job.per_file_results_json or "[]")
        job.status           = "awaiting_summary"
        job.progress_current = job.progress_total or 0
        job.status_detail    = f"Analiza par zakończona — {sum(len(r.get('changes',[])) for r in per_file_results)} zmian w {len(per_file_results)} parach"
        db.session.commit()
    return jsonify({"ok": True})


@bp.route("/job/<int:job_id>/run-pair-batch", methods=["POST"])
def run_pair_batch(job_id):
    """Process one batch of N sections for a file pair (browser-driven, avoids Apache timeout)."""
    job = ComparisonJob.query.get_or_404(job_id)

    db.session.refresh(job)
    if job.status == "cancelled":
        return jsonify({"ok": False, "cancelled": True, "error": "Porównanie zostało anulowane"})

    settings = db.session.get(AppSettings, 1)
    if not settings or not settings.gemini_api_key:
        return jsonify({"ok": False, "error": "Brak klucza Gemini API w ustawieniach"}), 400

    data = request.get_json() or {}
    pair_idx       = data.get("pair_idx", 0)
    section_offset = data.get("section_offset", 0)

    mappings = json.loads(job.file_mappings_json or "[]")
    if pair_idx >= len(mappings):
        return jsonify({"ok": False, "error": f"Nieprawidłowy indeks pary: {pair_idx}"}), 400

    mapping = mappings[pair_idx]
    doc_old = db.session.get(Document, mapping["old_doc_id"])
    doc_new = db.session.get(Document, mapping["new_doc_id"])
    if not doc_old or not doc_new:
        return jsonify({"ok": False, "pair_idx": pair_idx, "error": "Dokument nie istnieje"}), 404

    if not job.started_at:
        job.started_at = datetime.utcnow()
    job.status = "comparing"
    job.pair_lock_at = datetime.utcnow()
    pair_prefix = f"Para {pair_idx + 1}/{len(mappings)}"
    job.status_detail = f"{pair_prefix}: ekstrakcja struktury"
    db.session.commit()

    log.debug("run_pair_batch START  job=%d  para=%d  offset=%d  stary=%s  nowy=%s",
              job_id, pair_idx + 1, section_offset,
              doc_old.original_name, doc_new.original_name)

    try:
        struct_old, struct_new, all_keys, t_in_struct, t_out_struct = get_pair_structures(
            doc_old, doc_new, settings
        )

        # Renew lock after structure extraction (can take several minutes for large docs)
        job.pair_lock_at  = datetime.utcnow()
        job.tokens_input  = (job.tokens_input  or 0) + t_in_struct
        job.tokens_output = (job.tokens_output or 0) + t_out_struct

        import math as _math
        sections_total = len(all_keys)
        batch_end      = min(section_offset + BATCH_SIZE, sections_total)
        batch_keys     = all_keys[section_offset:batch_end]
        batch_num      = section_offset // BATCH_SIZE + 1
        total_batches  = _math.ceil(sections_total / BATCH_SIZE) if sections_total > 0 else 1

        sekcje_old = struct_old.get("sekcje", {})
        sekcje_new = struct_new.get("sekcje", {})

        pct_start = round(section_offset / sections_total * 100) if sections_total > 0 else 0
        job.status_detail = f"sekcja {section_offset + 1}/{sections_total} ({pct_start}%)"
        db.session.commit()

        call_fn, batch_tokens = make_gemini_caller(settings)

        def _on_progress(stage):
            job.status_detail = stage
            job.pair_lock_at = datetime.utcnow()
            try:
                db.session.commit()
            except Exception:
                pass

        changes_batch = compare_sections_batch(
            sekcje_old, sekcje_new, batch_keys,
            job.label_old, job.label_new,
            call_fn, settings, _on_progress,
            section_offset=section_offset,
            sections_total=sections_total,
            skip_redactional=bool(job.skip_redactional),
        )

        job.tokens_input  = (job.tokens_input  or 0) + batch_tokens["in"]
        job.tokens_output = (job.tokens_output or 0) + batch_tokens["out"]
        _update_cost(job)

        next_offset = batch_end
        done = next_offset >= sections_total

        log.debug("run_pair_batch OK  job=%d  para=%d  offset=%d→%d/%d  batch=%d/%d  zmian=%d  done=%s",
                  job_id, pair_idx + 1, section_offset, next_offset, sections_total,
                  batch_num, total_batches, len(changes_batch), done)

        db.session.commit()

        return jsonify({
            "ok":            True,
            "pair_idx":      pair_idx,
            "sections_total": sections_total,
            "next_offset":   next_offset,
            "done":          done,
            "changes_batch": changes_batch,
            "batch_num":     batch_num,
            "total_batches": total_batches,
        })

    except Exception as exc:
        log.error("run_pair_batch BŁĄD  job=%d  para=%d  offset=%d  %s: %s",
                  job_id, pair_idx + 1, section_offset,
                  type(exc).__name__, exc, exc_info=True)
        try:
            job.status        = "error"
            job.error_message = str(exc)[:1000]
            db.session.commit()
        except Exception as db_err:
            log.error("run_pair_batch DB commit nieudany: %s", db_err)
            try: db.session.rollback()
            except Exception: pass
        return jsonify({"ok": False, "pair_idx": pair_idx, "error": str(exc)}), 500


@bp.route("/job/<int:job_id>/finalize-pair", methods=["POST"])
def finalize_pair(job_id):
    """Generate per-file summary from all accumulated changes, then persist the result."""
    job = ComparisonJob.query.get_or_404(job_id)

    settings = db.session.get(AppSettings, 1)
    if not settings or not settings.gemini_api_key:
        return jsonify({"ok": False, "error": "Brak klucza Gemini API w ustawieniach"}), 400

    data = request.get_json() or {}
    pair_idx    = data.get("pair_idx", 0)
    all_changes = data.get("changes", [])

    mappings = json.loads(job.file_mappings_json or "[]")
    if pair_idx >= len(mappings):
        return jsonify({"ok": False, "error": f"Nieprawidłowy indeks pary: {pair_idx}"}), 400

    mapping = mappings[pair_idx]
    doc_old = db.session.get(Document, mapping["old_doc_id"])
    doc_new = db.session.get(Document, mapping["new_doc_id"])
    if not doc_old or not doc_new:
        return jsonify({"ok": False, "pair_idx": pair_idx, "error": "Dokument nie istnieje"}), 404

    pair_prefix = f"Para {pair_idx + 1}/{len(mappings)}"
    job.status_detail = f"{pair_prefix}: generuję podsumowanie pliku..."
    job.pair_lock_at = datetime.utcnow()
    db.session.commit()

    log.debug("finalize_pair START  job=%d  para=%d  zmian=%d",
              job_id, pair_idx + 1, len(all_changes))

    try:
        summary_text, t_in, t_out = generate_pair_summary(
            all_changes,
            job.label_old, job.label_new,
            job.competition_name, settings,
        )

        result = {
            "pair_idx":    pair_idx,
            "old_doc_id":  doc_old.id,
            "new_doc_id":  doc_new.id,
            "old_name":    doc_old.original_name or mapping.get("old_name", ""),
            "new_name":    doc_new.original_name or mapping.get("new_name", ""),
            "changes":     all_changes,
            "summary":     summary_text,
            "tokens_in":   t_in,
            "tokens_out":  t_out,
            "compared_at": datetime.utcnow().strftime("%d.%m.%Y %H:%M"),
        }

        per_file_results = json.loads(job.per_file_results_json or "[]")
        per_file_results = [r for r in per_file_results if r.get("pair_idx") != pair_idx]
        per_file_results.append(result)
        job.per_file_results_json = json.dumps(per_file_results, ensure_ascii=False)

        all_c = [c for r in per_file_results for c in r.get("changes", [])]
        job.changes_json     = json.dumps(all_c, ensure_ascii=False)
        job.progress_current = pair_idx + 1
        job.tokens_input     = (job.tokens_input  or 0) + t_in
        job.tokens_output    = (job.tokens_output or 0) + t_out
        job.error_message    = None
        _update_cost(job)
        db.session.commit()

        log.debug("finalize_pair OK  job=%d  para=%d  zmian=%d",
                  job_id, pair_idx + 1, len(all_changes))

        import markdown as md_lib
        summary_html = md_lib.markdown(summary_text, extensions=["extra"]) if summary_text else ""

        return jsonify({
            "ok":           True,
            "pair_idx":     pair_idx,
            "old_name":     result["old_name"],
            "new_name":     result["new_name"],
            "changes":      all_changes,
            "summary":      summary_text,
            "summary_html": summary_html,
        })

    except Exception as exc:
        log.error("finalize_pair BŁĄD  job=%d  para=%d  %s: %s",
                  job_id, pair_idx + 1, type(exc).__name__, exc, exc_info=True)
        try:
            job.status        = "error"
            job.error_message = str(exc)[:1000]
            db.session.commit()
        except Exception as db_err:
            log.error("finalize_pair DB commit nieudany: %s", db_err)
            try: db.session.rollback()
            except Exception: pass
        return jsonify({"ok": False, "pair_idx": pair_idx, "error": str(exc)}), 500


@bp.route("/job/<int:job_id>/generate-summary", methods=["POST"])
def generate_edition_summary(job_id):
    """Generate the edition-wide executive summary from completed per-file results."""
    is_ajax = (request.is_json
               or request.headers.get("X-Requested-With") == "XMLHttpRequest"
               or request.accept_mimetypes.best == "application/json")

    job = ComparisonJob.query.get_or_404(job_id)
    settings = db.session.get(AppSettings, 1)

    if not settings or not settings.gemini_api_key:
        if is_ajax:
            return jsonify({"ok": False, "error": "Brak klucza Gemini API w ustawieniach"}), 400
        flash("Brak klucza Gemini API w ustawieniach.", "error")
        return redirect(url_for("comparison.job_status", job_id=job_id))

    per_file_results = json.loads(job.per_file_results_json or "[]")
    if not per_file_results:
        if is_ajax:
            return jsonify({"ok": False, "error": "Brak wyników do podsumowania"}), 400
        flash("Brak wyników par plików — uruchom najpierw analizę.", "warning")
        return redirect(url_for("comparison.job_status", job_id=job_id))

    job.status        = "summarizing"
    job.status_detail = "Generuję podsumowanie całej edycji..."
    db.session.commit()

    try:
        summary_text, t_in, t_out = generate_edition_summary_text(per_file_results, job, settings)

        job.edition_summary  = summary_text
        job.status           = "done"
        job.status_detail    = f"Analiza zakończona — {sum(len(r.get('changes',[])) for r in per_file_results)} zmian w {len(per_file_results)} parach"
        job.finished_at      = datetime.utcnow()
        job.tokens_input     = (job.tokens_input  or 0) + t_in
        job.tokens_output    = (job.tokens_output or 0) + t_out
        _update_cost(job)
        db.session.commit()

        _promote_next_queued()

        if is_ajax:
            import markdown as md_lib
            summary_html = md_lib.markdown(summary_text, extensions=["extra"])
            return jsonify({"ok": True, "summary_html": summary_html})
        flash("Podsumowanie edycji wygenerowane.", "success")
        return redirect(url_for("comparison.job_status", job_id=job_id))
    except Exception as exc:
        log.error("generate_edition_summary BŁĄD  job=%d  %s: %s",
                  job_id, type(exc).__name__, exc, exc_info=True)
        job.status        = "error"
        job.error_message = str(exc)[:1000]
        db.session.commit()
        _promote_next_queued()
        if is_ajax:
            return jsonify({"ok": False, "error": str(exc)}), 500
        flash(f"Błąd generowania podsumowania: {exc}", "error")
        return redirect(url_for("comparison.job_status", job_id=job_id))


def _update_cost(job):
    model = job.gemini_model_used or "gemini-2.5-flash"
    price = PRICING.get(model, {"input": 0.30, "output": 2.50})
    t_in  = job.tokens_input  or 0
    t_out = job.tokens_output or 0
    job.estimated_cost_usd = round(
        (t_in / 1_000_000 * price["input"]) + (t_out / 1_000_000 * price["output"]), 6
    )


def _promote_next_queued():
    """Promote the oldest queued job to pending so it can start."""
    next_job = ComparisonJob.query.filter_by(status="queued").order_by(ComparisonJob.created_at).first()
    if next_job:
        next_job.status = "pending"
        db.session.commit()


# ── Result / status pages ──────────────────────────────────────────────────

@bp.route("/job/<int:job_id>")
def job_status(job_id):
    import markdown as md_lib
    job = ComparisonJob.query.get_or_404(job_id)

    per_file_results = []
    all_changes = []

    if job.per_file_results_json:
        per_file_results = json.loads(job.per_file_results_json)
        for r in per_file_results:
            all_changes.extend(r.get("changes", []))
            if r.get("summary"):
                r["summary_html"] = md_lib.markdown(r["summary"], extensions=["extra"])

    edition_summary_html = ""
    if job.edition_summary:
        edition_summary_html = md_lib.markdown(job.edition_summary, extensions=["extra"])

    all_mappings = json.loads(job.file_mappings_json or "[]")
    # Use pair_idx for deduplication (supports same file pair appearing multiple times)
    done_indices = {r["pair_idx"] for r in per_file_results if "pair_idx" in r}
    # Backward compat: old records without pair_idx use doc_id pair
    done_doc_pairs = {(r["old_doc_id"], r["new_doc_id"]) for r in per_file_results if "pair_idx" not in r}
    pending_pairs = [
        {"idx": i, "old_name": m.get("old_name", ""), "new_name": m.get("new_name", "")}
        for i, m in enumerate(all_mappings)
        if i not in done_indices and (m["old_doc_id"], m["new_doc_id"]) not in done_doc_pairs
    ]
    used_old_ids = [m["old_doc_id"] for m in all_mappings]
    used_new_ids = [m["new_doc_id"] for m in all_mappings]

    pair_lock_age_secs = None
    if job.pair_lock_at:
        pair_lock_age_secs = int((datetime.utcnow() - job.pair_lock_at).total_seconds())

    return render_template(
        "comparison/result.html",
        job=job,
        changes=all_changes,
        per_file_results=per_file_results,
        edition_summary_html=edition_summary_html,
        pending_pairs=pending_pairs,
        used_old_ids=used_old_ids,
        used_new_ids=used_new_ids,
        pair_lock_age_secs=pair_lock_age_secs,
    )


@bp.route("/job/<int:job_id>/status-api")
def job_status_api(job_id):
    job = ComparisonJob.query.get_or_404(job_id)
    pair_lock_age_secs = None
    if job.pair_lock_at:
        pair_lock_age_secs = int((datetime.utcnow() - job.pair_lock_at).total_seconds())
    return jsonify({
        "status":             job.status,
        "status_detail":      job.status_detail or "",
        "progress_current":   job.progress_current,
        "progress_total":     job.progress_total,
        "error_message":      job.error_message,
        "pair_lock_age_secs": pair_lock_age_secs,
    })


@bp.route("/job/<int:job_id>/pair-result/<int:pair_idx>")
def pair_result(job_id, pair_idx):
    """Return the saved result for one pair (or done=false if not yet finished)."""
    import markdown as md_lib
    job = ComparisonJob.query.get_or_404(job_id)
    mappings = json.loads(job.file_mappings_json or "[]")
    if pair_idx >= len(mappings):
        return jsonify({"done": False})
    m = mappings[pair_idx]
    per_file_results = json.loads(job.per_file_results_json or "[]")
    for r in per_file_results:
        # Match by pair_idx (new records) or by doc_id pair (old records)
        matches = (r.get("pair_idx") == pair_idx) or (
            "pair_idx" not in r
            and r.get("old_doc_id") == m["old_doc_id"]
            and r.get("new_doc_id") == m["new_doc_id"]
        )
        if matches:
            summary_html = md_lib.markdown(r.get("summary", ""), extensions=["extra"]) if r.get("summary") else ""
            return jsonify({
                "done":         True,
                "skipped":      r.get("skipped", False),
                "old_name":     r.get("old_name", ""),
                "new_name":     r.get("new_name", ""),
                "changes":      r.get("changes", []),
                "summary":      r.get("summary", ""),
                "summary_html": summary_html,
            })
    return jsonify({"done": False})


# ── Skip pair ─────────────────────────────────────────────────────────────

@bp.route("/job/<int:job_id>/reset-pair", methods=["POST"])
def reset_pair(job_id):
    """Remove a completed pair result so it can be re-run."""
    job = ComparisonJob.query.get_or_404(job_id)
    data = request.get_json() or {}
    pair_idx = data.get("pair_idx")
    if pair_idx is None:
        return jsonify({"ok": False, "error": "pair_idx required"}), 400

    per_file_results = json.loads(job.per_file_results_json or "[]")
    per_file_results = [r for r in per_file_results if r.get("pair_idx") != pair_idx]
    job.per_file_results_json = json.dumps(per_file_results, ensure_ascii=False)

    # Clear edition summary — it's now stale
    job.edition_summary = None

    # Allow pair to be processed again
    job.pair_lock_at = None
    if job.status in ("done", "awaiting_summary"):
        job.status = "comparing"
        job.finished_at = None

    # Provide mapping info so the browser can queue the pair
    all_mappings = json.loads(job.file_mappings_json or "[]")
    mapping = all_mappings[pair_idx] if pair_idx < len(all_mappings) else {}

    db.session.commit()
    return jsonify({
        "ok":        True,
        "pair_idx":  pair_idx,
        "old_doc_id": mapping.get("old_doc_id"),
        "new_doc_id": mapping.get("new_doc_id"),
        "old_name":   mapping.get("old_name", ""),
        "new_name":   mapping.get("new_name", ""),
    })


@bp.route("/job/<int:job_id>/rename", methods=["POST"])
def rename_job(job_id):
    job = ComparisonJob.query.get_or_404(job_id)
    data = request.get_json() or {}
    label = (data.get("label") or "").strip()
    job.job_label = label or None
    db.session.commit()
    return jsonify({"ok": True, "label": job.job_label or job.competition_name or f"Porównanie #{job.id}"})


@bp.route("/job/<int:job_id>/add-pair", methods=["POST"])
def add_pair(job_id):
    """Append a new file pair to an existing comparison job."""
    job = ComparisonJob.query.get_or_404(job_id)

    if job.status in ("cancelled", "queued"):
        return jsonify({"ok": False, "error": "Nie można dodać pary do anulowanego lub oczekującego zadania"}), 400

    data = request.get_json() or {}
    old_doc_id = data.get("old_doc_id")
    new_doc_id = data.get("new_doc_id")

    if not old_doc_id or not new_doc_id:
        return jsonify({"ok": False, "error": "Brak identyfikatorów dokumentów"}), 400

    doc_old = db.session.get(Document, old_doc_id)
    doc_new = db.session.get(Document, new_doc_id)

    if not doc_old or not doc_new:
        return jsonify({"ok": False, "error": "Dokument nie istnieje"}), 404

    mappings = json.loads(job.file_mappings_json or "[]")
    new_idx = len(mappings)
    mappings.append({
        "old_doc_id": old_doc_id,
        "new_doc_id": new_doc_id,
        "old_name":   doc_old.original_name,
        "new_name":   doc_new.original_name,
    })

    job.file_mappings_json = json.dumps(mappings, ensure_ascii=False)
    job.progress_total     = (job.progress_total or 0) + 1
    db.session.commit()

    log.debug("add_pair OK  job=%d  pair_idx=%d  stary=%s  nowy=%s",
              job_id, new_idx, doc_old.original_name, doc_new.original_name)

    return jsonify({
        "ok":       True,
        "pair_idx": new_idx,
        "old_name": doc_old.original_name,
        "new_name": doc_new.original_name,
    })


@bp.route("/job/<int:job_id>/skip-pair", methods=["POST"])
def skip_pair(job_id):
    """Mark a file pair as skipped (no Gemini call, empty changes)."""
    job = ComparisonJob.query.get_or_404(job_id)
    data = request.get_json() or {}
    pair_idx = data.get("pair_idx", 0)

    mappings = json.loads(job.file_mappings_json or "[]")
    if pair_idx >= len(mappings):
        return jsonify({"ok": False, "error": "Nieprawidłowy indeks pary"}), 400

    mapping = mappings[pair_idx]
    old_doc_id = mapping["old_doc_id"]
    new_doc_id = mapping["new_doc_id"]

    per_file_results = json.loads(job.per_file_results_json or "[]")
    per_file_results = [r for r in per_file_results if r.get("pair_idx") != pair_idx]
    per_file_results.append({
        "pair_idx":    pair_idx,
        "old_doc_id":  old_doc_id,
        "new_doc_id":  new_doc_id,
        "old_name":    mapping.get("old_name", ""),
        "new_name":    mapping.get("new_name", ""),
        "changes":     [],
        "summary":     "",
        "skipped":     True,
        "compared_at": datetime.utcnow().strftime("%d.%m.%Y %H:%M"),
    })
    job.per_file_results_json = json.dumps(per_file_results, ensure_ascii=False)
    db.session.commit()

    return jsonify({
        "ok":       True,
        "pair_idx": pair_idx,
        "old_name": mapping.get("old_name", ""),
        "new_name": mapping.get("new_name", ""),
    })


# ── Excel export ───────────────────────────────────────────────────────────

_BOTTLE     = "0019A6"
_WAGA_FILL  = {"KRYTYCZNA": "FFCCCC", "WYSOKA": "FFE5CC", "SREDNIA": "FFFFCC", "NISKA": "E5FFE5"}

_ILLEGAL_XML = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]')
_XL_MAX_LEN  = 32000  # Excel hard limit is 32 767; leave a small buffer


def _xl_val(v):
    """Strip XML-illegal control characters and cap length for Excel cell safety."""
    if not isinstance(v, str):
        return v
    v = _ILLEGAL_XML.sub('', v)
    if len(v) > _XL_MAX_LEN:
        v = v[:_XL_MAX_LEN] + '…'
    return v


def _excel_filename(job, suffix=""):
    import re
    dt = job.created_at.strftime("%Y-%m-%d %H-%M") if job.created_at else "export"
    parts = [job.competition_name or "konkurs", job.label_new or "edycja", dt]
    if suffix:
        parts.append(suffix)
    slug = "-".join(re.sub(r"[^\w\s\-]", "", p or "").strip() for p in parts)
    return secure_filename(f"Rejestr zmian-{slug}.xlsx")


def _safe_sheet_name(name):
    import re
    return re.sub(r"[\\/*?:\[\]]", "_", name or "Arkusz")[:31]


def _xl_col_widths(ws, widths):
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_summary_sheet(ws, job, per_file_results):
    from openpyxl.styles import Alignment, Font, PatternFill

    NCOLS = 7
    ws["A1"] = _xl_val(job.competition_name or "Porównanie edycji")
    ws["A1"].font = Font(bold=True, size=14, color=_BOTTLE)
    ws["A2"] = _xl_val(f"{job.label_old or '?'}  →  {job.label_new or '?'}")
    ws["A2"].font = Font(size=14, color="444444")
    ws["A3"] = f"Data: {job.created_at.strftime('%d.%m.%Y %H:%M') if job.created_at else '—'}"
    ws["A3"].font = Font(size=14, color="888888")
    for r in [1, 2, 3]:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)

    ROW_HDR = 5
    for col, label in enumerate(
        ["Starsza edycja (plik)", "Nowsza edycja (plik)",
         "Zmian", "Krytyczne", "Wysokie", "Średnie", "Niskie"], 1
    ):
        c = ws.cell(row=ROW_HDR, column=col, value=label)
        c.font = Font(bold=True, size=14, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=_BOTTLE)
        c.alignment = Alignment(wrap_text=True, horizontal="center" if col > 2 else "left")

    row = ROW_HDR
    active_pfr = [p for p in per_file_results if not p.get("skipped")]
    for pfr in active_pfr:
        row += 1
        ch     = pfr.get("changes", [])
        n_tot  = len(ch)
        n_crit = sum(1 for c in ch if c.get("waga") == "KRYTYCZNA")
        n_high = sum(1 for c in ch if c.get("waga") == "WYSOKA")
        n_med  = sum(1 for c in ch if c.get("waga") == "SREDNIA")
        n_low  = sum(1 for c in ch if c.get("waga") == "NISKA")
        for col, v in enumerate(
            [_xl_val(pfr.get("old_name","")), _xl_val(pfr.get("new_name","")),
             n_tot, n_crit, n_high, n_med, n_low], 1
        ):
            cell = ws.cell(row=row, column=col, value=v)
            cell.alignment = Alignment(wrap_text=True,
                                       horizontal="center" if col > 2 else "left")
        if n_crit > 0:
            ws.cell(row=row, column=4).fill = PatternFill("solid", fgColor="FFCCCC")

    # Total row
    row += 1
    all_ch = [c for p in active_pfr for c in p.get("changes", [])]
    for col, v in enumerate(
        ["ŁĄCZNIE", "", len(all_ch),
         sum(1 for c in all_ch if c.get("waga") == "KRYTYCZNA"),
         sum(1 for c in all_ch if c.get("waga") == "WYSOKA"),
         sum(1 for c in all_ch if c.get("waga") == "SREDNIA"),
         sum(1 for c in all_ch if c.get("waga") == "NISKA")], 1
    ):
        cell = ws.cell(row=row, column=col, value=v)
        cell.font = Font(bold=True, size=14)
        cell.fill = PatternFill("solid", fgColor="EEEEEE")
        cell.alignment = Alignment(horizontal="center" if col > 2 else "left")

    # AI edition summary
    if job.edition_summary:
        row += 2
        lbl = ws.cell(row=row, column=1, value="Podsumowanie AI całej edycji:")
        lbl.font = Font(bold=True, size=14, color=_BOTTLE)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
        row += 1
        cell = ws.cell(row=row, column=1, value=_xl_val(job.edition_summary))
        cell.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
        ws.row_dimensions[row].height = min(400, max(60, len(job.edition_summary) // 8))

    _xl_col_widths(ws, [40, 40, 10, 12, 10, 10, 10])


def _write_pair_sheet(ws, pfr, job):
    from openpyxl.styles import Alignment, Font, PatternFill

    old_name = _xl_val(pfr.get("old_name", "?"))
    new_name = _xl_val(pfr.get("new_name", "?"))
    summary  = _xl_val(pfr.get("summary", ""))
    changes  = pfr.get("changes", [])

    ws["A1"] = old_name
    ws["A1"].font = Font(bold=True, size=14, color=_BOTTLE)
    ws["A2"] = f"→  {new_name}"
    ws["A2"].font = Font(size=14, color="555555")
    for r in [1, 2]:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)

    row = 3

    if summary:
        row += 1
        lbl = ws.cell(row=row, column=1, value="Podsumowanie zmian w pliku:")
        lbl.font = Font(bold=True, size=14, color=_BOTTLE)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
        cell = ws.cell(row=row, column=1, value=summary)
        cell.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.row_dimensions[row].height = min(400, max(60, len(summary) // 8))
        row += 2

    if not changes:
        ws.cell(row=row, column=1,
                value="Brak zmian." if not pfr.get("skipped") else "Para pominięta.").font = \
            Font(italic=True, size=14, color="888888")
        _xl_col_widths(ws, [20, 20, 12, 50, 50, 60])
        return

    tbl_row = row
    for col, h in enumerate(
        ["Sekcja", "Typ zmiany", "Waga",
         _xl_val(f"Zapis — {job.label_old}"), _xl_val(f"Zapis — {job.label_new}"),
         "Komentarz biznesowy"], 1
    ):
        c = ws.cell(row=tbl_row, column=col, value=h)
        c.font = Font(bold=True, size=14, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=_BOTTLE)
        c.alignment = Alignment(wrap_text=True)

    for change in changes:
        row += 1
        waga = _xl_val(change.get("waga", "NISKA"))
        fill = PatternFill("solid", fgColor=_WAGA_FILL.get(waga, "FFFFFF"))
        for col, v in enumerate(
            [_xl_val(change.get("sekcja","")),
             _xl_val(change.get("typ_zmiany","")),
             waga,
             _xl_val(change.get("zapis_stary","")),
             _xl_val(change.get("zapis_nowy","")),
             _xl_val(change.get("komentarz_biznesowy",""))], 1
        ):
            cell = ws.cell(row=row, column=col, value=v)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True)

    for r in ws.iter_rows(min_row=tbl_row + 1):
        ws.row_dimensions[r[0].row].height = 80
    _xl_col_widths(ws, [20, 20, 12, 50, 50, 60])


def _send_excel(wb, filename):
    """Serialize workbook to bytes and return as a response (avoids BytesIO.fileno issue)."""
    from flask import make_response as _make_response
    buf = io.BytesIO()
    wb.save(buf)
    resp = _make_response(buf.getvalue())
    resp.headers["Content-Type"] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _send_csv(content_str, filename):
    """Return CSV string as a download response with UTF-8 BOM for Excel compatibility."""
    from flask import make_response as _make_response
    resp = _make_response(content_str.encode("utf-8-sig"))
    resp.headers["Content-Type"] = "text/csv; charset=utf-8"
    resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _csv_filename(job, suffix=""):
    import re
    dt = job.created_at.strftime("%Y-%m-%d %H-%M") if job.created_at else "export"
    parts = [job.competition_name or "konkurs", job.label_new or "edycja", dt]
    if suffix:
        parts.append(suffix)
    slug = "-".join(re.sub(r"[^\w\s\-]", "", p or "").strip() for p in parts)
    return f"Rejestr zmian-{slug}.csv"


def _build_pair_csv(pfr, job):
    import csv as _csv
    import io as _sio
    changes = pfr.get("changes", [])
    buf = _sio.StringIO()
    w = _csv.writer(buf)
    w.writerow([
        "Sekcja", "Typ zmiany", "Waga",
        f"Zapis — {job.label_old}", f"Zapis — {job.label_new}",
        "Komentarz biznesowy",
    ])
    for ch in changes:
        w.writerow([
            ch.get("sekcja", ""),
            ch.get("typ_zmiany", ""),
            ch.get("waga", ""),
            ch.get("zapis_stary", ""),
            ch.get("zapis_nowy", ""),
            ch.get("komentarz_biznesowy", ""),
        ])
    return buf.getvalue()


def _build_all_csv(per_file_results, job):
    import csv as _csv
    import io as _sio
    active = [p for p in per_file_results if not p.get("skipped")]
    buf = _sio.StringIO()
    w = _csv.writer(buf)
    w.writerow([
        "Para (stary plik)", "Para (nowy plik)",
        "Sekcja", "Typ zmiany", "Waga",
        f"Zapis — {job.label_old}", f"Zapis — {job.label_new}",
        "Komentarz biznesowy",
    ])
    for pfr in active:
        for ch in pfr.get("changes", []):
            w.writerow([
                pfr.get("old_name", ""),
                pfr.get("new_name", ""),
                ch.get("sekcja", ""),
                ch.get("typ_zmiany", ""),
                ch.get("waga", ""),
                ch.get("zapis_stary", ""),
                ch.get("zapis_nowy", ""),
                ch.get("komentarz_biznesowy", ""),
            ])
    return buf.getvalue()


@bp.route("/job/<int:job_id>/download-excel")
def download_excel(job_id):
    import openpyxl
    job = ComparisonJob.query.get_or_404(job_id)
    per_file_results = json.loads(job.per_file_results_json or "[]")

    if not per_file_results:
        flash("Brak wyników do eksportu.", "warning")
        return redirect(url_for("comparison.job_status", job_id=job_id))

    wb = openpyxl.Workbook()
    ws_sum = wb.active
    ws_sum.title = "Podsumowanie"
    _write_summary_sheet(ws_sum, job, per_file_results)

    for pfr in per_file_results:
        if pfr.get("skipped"):
            continue
        ws = wb.create_sheet(_safe_sheet_name(pfr.get("old_name", "Para")))
        _write_pair_sheet(ws, pfr, job)

    return _send_excel(wb, _excel_filename(job))


@bp.route("/job/<int:job_id>/download-csv")
def download_csv(job_id):
    job = ComparisonJob.query.get_or_404(job_id)
    per_file_results = json.loads(job.per_file_results_json or "[]")

    if not per_file_results:
        flash("Brak wyników do eksportu.", "warning")
        return redirect(url_for("comparison.job_status", job_id=job_id))

    return _send_csv(_build_all_csv(per_file_results, job), _csv_filename(job))


@bp.route("/job/<int:job_id>/pair/<int:pair_idx>/download-excel")
def download_pair_excel(job_id, pair_idx):
    import openpyxl
    job = ComparisonJob.query.get_or_404(job_id)
    per_file_results = json.loads(job.per_file_results_json or "[]")

    if pair_idx >= len(per_file_results):
        flash("Nie znaleziono wyników tej pary.", "warning")
        return redirect(url_for("comparison.job_status", job_id=job_id))

    pfr = per_file_results[pair_idx]
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = _safe_sheet_name(pfr.get("old_name", "Para"))
    _write_pair_sheet(ws, pfr, job)

    return _send_excel(wb, _excel_filename(job, suffix=pfr.get("old_name", f"para-{pair_idx + 1}")))


@bp.route("/job/<int:job_id>/pair/<int:pair_idx>/download-csv")
def download_pair_csv(job_id, pair_idx):
    job = ComparisonJob.query.get_or_404(job_id)
    per_file_results = json.loads(job.per_file_results_json or "[]")

    if pair_idx >= len(per_file_results):
        flash("Nie znaleziono wyników tej pary.", "warning")
        return redirect(url_for("comparison.job_status", job_id=job_id))

    pfr = per_file_results[pair_idx]
    return _send_csv(
        _build_pair_csv(pfr, job),
        _csv_filename(job, suffix=pfr.get("old_name", f"para-{pair_idx + 1}")),
    )


@bp.route("/job/<int:job_id>/cancel", methods=["POST"])
def cancel_job(job_id):
    job = ComparisonJob.query.get_or_404(job_id)
    cancellable = {"queued", "pending", "comparing", "extracting", "chunking", "awaiting_summary", "summarizing"}
    if job.status not in cancellable:
        if request.is_json or request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({"ok": False, "error": f"Nie można anulować statusu '{job.status}'"}), 400
        flash(f"Nie można anulować porównania o statusie '{job.status}'.", "warning")
        return redirect(url_for("comparison.index"))

    job.status      = "cancelled"
    job.finished_at = datetime.utcnow()
    db.session.commit()

    _promote_next_queued()

    if request.is_json or request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return jsonify({"ok": True})
    flash("Porównanie anulowane.", "info")
    return redirect(url_for("comparison.index"))


@bp.route("/job/<int:job_id>/delete", methods=["POST"])
def delete_job(job_id):
    job = ComparisonJob.query.get_or_404(job_id)
    db.session.delete(job)
    db.session.commit()
    flash("Porównanie usunięte.", "success")
    return redirect(url_for("comparison.index"))
